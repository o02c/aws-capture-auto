from pathlib import Path
from pydantic import BaseModel, Field, validator
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from typing import Optional
from shutil import copy2

class ImageToExcelConfig(BaseModel):
    """画像をExcelに貼り付けるための設定を格納するクラス"""
    image_path: str = Field(..., description="画像ファイルのパス")
    sheet_name: str | int = Field(..., description="貼り付けるシート名またはインデックス（0始まり）")
    cell: str = Field(..., description="貼り付けるセル位置 (例: 'A1')")
    width_cm: float = Field(..., gt=0, description="貼り付ける画像の幅 (cm)")
    height_cm: float = Field(..., gt=0, description="貼り付ける画像の高さ (cm)")

    @validator('image_path')
    def validate_image_path(cls, v):
        """画像ファイルのパスを検証"""
        path = Path(v)
        if not path.exists():
            raise ValueError(f"画像ファイルが存在しません: {v}")
        if path.suffix.lower() not in ['.png', '.jpg', '.jpeg', '.gif', '.bmp']:
            raise ValueError(f"サポートされていない画像形式です: {path.suffix}")
        return str(path)

    @validator('cell')
    def validate_cell(cls, v):
        """セル位置の形式を検証"""
        # 文字列が空でないことを確認
        if not v:
            raise ValueError("セル位置が指定されていません")
        
        # アルファベット部分と数字部分に分割
        alpha_part = ""
        digit_part = ""
        
        for char in v:
            if char.isalpha():
                if digit_part:  # 数字の後にアルファベットが来た場合
                    raise ValueError(f"無効なセル位置です: {v}")
                alpha_part += char.upper()
            elif char.isdigit():
                digit_part += char
            else:
                raise ValueError(f"無効なセル位置です: {v}")
        
        # 両方の部分が存在することを確認
        if not alpha_part or not digit_part:
            raise ValueError(f"無効なセル位置です: {v}")
            
        return v

    @validator('sheet_name')
    def validate_sheet_name(cls, v):
        """シート名または番号を検証"""
        if isinstance(v, int) and v < 0:
            raise ValueError("シート番号は0以上の値を指定してください")
        return v

    def to_pixels(self) -> tuple[float, float]:
        """cmをピクセルに変換"""
        PIXELS_PER_CM = 37.795275591  # 1cmあたりのピクセル数
        width_px = self.width_cm * PIXELS_PER_CM
        height_px = self.height_cm * PIXELS_PER_CM
        return width_px, height_px

def insert_image_to_excel(
    configs: list[ImageToExcelConfig], 
    input_excel: Path, 
    output_excel: Optional[Path] = None
) -> Path:
    """画像をExcelの指定位置に貼り付ける
    
    Args:
        configs (list[ImageToExcelConfig]): 画像貼り付けの設定のリスト
        input_excel (Path): 入力Excelファイルのパス
        output_excel (Optional[Path]): 出力Excelファイルのパス。
            Noneの場合は入力ファイル名に '_with_images' を追加
    
    Returns:
        Path: 出力されたExcelファイルのパス
    """
    if not input_excel.exists():
        raise ValueError(f"入力Excelファイルが存在しません: {input_excel}")

    # 出力パスが指定されていない場合は入力ファイル名に基づいて生成
    if output_excel is None:
        output_excel = input_excel.parent / f"{input_excel.stem}_with_images{input_excel.suffix}"
    
    # 入力ファイルを出力先にコピー
    copy2(input_excel, output_excel)
    
    # Excelファイルを読み込む
    wb = load_workbook(output_excel)
    
    for config in configs:
        # シートを取得
        if isinstance(config.sheet_name, int):
            if config.sheet_name >= len(wb.sheetnames):
                raise ValueError(f"シート番号 {config.sheet_name} は範囲外です")
            sheet_name = wb.sheetnames[config.sheet_name]
        else:
            sheet_name = config.sheet_name
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"シート '{sheet_name}' が見つかりません")
        
        ws = wb[sheet_name]
        
        # 画像をExcel用に変換
        excel_img = Image(config.image_path)
        
        # アスペクト比を計算
        aspect_ratio = excel_img.width / excel_img.height
        
        # cmをピクセルに変換
        width_px, height_px = config.to_pixels()
        
        # 指定されたサイズに合わせて画像をリサイズ
        if width_px / height_px > aspect_ratio:
            # 高さに合わせてリサイズ
            new_width = int(height_px * aspect_ratio)
            new_height = int(height_px)
        else:
            # 幅に合わせてリサイズ
            new_width = int(width_px)
            new_height = int(width_px / aspect_ratio)
        
        excel_img.width = new_width
        excel_img.height = new_height
        
        # 画像を貼り付ける
        ws.add_image(excel_img, config.cell)
    
    # 変更を保存
    wb.save(output_excel)
    
    return output_excel
